from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.core.exceptions import PermissionDenied
from apps.documentos.models import DocumentoInscripcion
from apps.notificaciones.services import enviar_ficha_postulante
from apps.pdf.services import generar_ficha_postulante_pdf
from apps.postulantes.models import Inscripcion
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

import logging

logger = logging.getLogger(__name__)


def usuario_interno_requerido(view_func):
    @login_required
    def wrapper(request, *args, **kwargs):
        if request.user.rol not in ["ADMINISTRADOR", "VALIDADOR", "CONSULTA"]:
            messages.error(request, "No tiene permisos para acceder al panel.")
            return redirect("nucleo:inicio")
        return view_func(request, *args, **kwargs)
    return wrapper


@usuario_interno_requerido
def dashboard(request):
    estado = request.GET.get("estado")

    qs = Inscripcion.objects.all()

    if estado:
        qs = qs.filter(estado=estado)

    context = {
        "total": Inscripcion.objects.count(),
        "registrado": Inscripcion.objects.filter(estado=Inscripcion.Estado.REGISTRADO).count(),
        "observado": Inscripcion.objects.filter(estado=Inscripcion.Estado.OBSERVADO).count(),
        "validado": Inscripcion.objects.filter(estado=Inscripcion.Estado.VALIDADO).count(),
        "inscripciones": qs.order_by("-fecha_creacion"),
        "estado_seleccionado": estado,
    }

    return render(request, "panel/dashboard.html", context)

@usuario_interno_requerido
def lista_postulantes(request):
    estado = request.GET.get("estado", "").strip()
    programa = request.GET.get("programa", "").strip()
    q = request.GET.get("q", "").strip()

    inscripciones = Inscripcion.objects.select_related(
        "postulante",
        "convocatoria",
        "modalidad",
        "primera_opcion_programa",
        "segunda_opcion_programa",
    ).order_by("-fecha_creacion")

    if estado:
        inscripciones = inscripciones.filter(estado=estado)

    if programa:
        inscripciones = inscripciones.filter(primera_opcion_programa__id=programa)

    if q:
        inscripciones = inscripciones.filter(
            postulante__nombres__icontains=q
        ) | inscripciones.filter(
            postulante__apellido_paterno__icontains=q
        ) | inscripciones.filter(
            postulante__apellido_materno__icontains=q
        ) | inscripciones.filter(
            postulante__numero_documento__icontains=q
        ) | inscripciones.filter(
            numero_inscripcion__icontains=q
        )

    context = {
        "inscripciones": inscripciones,
        "estado_actual": estado,
        "programa_actual": programa,
        "q": q,
    }
    return render(request, "panel/lista_postulantes.html", context)


@usuario_interno_requerido
def detalle_postulante(request, inscripcion_id):
    inscripcion = get_object_or_404(
        Inscripcion.objects.select_related(
            "postulante",
            "convocatoria",
            "modalidad",
            "primera_opcion_programa",
            "segunda_opcion_programa",
        ),
        id=inscripcion_id,
    )

    documentos = DocumentoInscripcion.objects.filter(
        inscripcion=inscripcion
    ).select_related("tipo_documento").order_by("tipo_documento__nombre")

    context = {
        "inscripcion": inscripcion,
        "postulante": inscripcion.postulante,
        "documentos": documentos,
    }
    return render(request, "panel/detalle_postulante.html", context)


@usuario_interno_requerido
def generar_ficha_completa(request, inscripcion_id):
    inscripcion = get_object_or_404(Inscripcion, id=inscripcion_id)

    pdf_bytes = generar_ficha_postulante_pdf(inscripcion)

    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = (
        f'inline; filename="ficha_completa_{inscripcion.numero_inscripcion}.pdf"'
    )
    return response


@usuario_interno_requerido

def reenviar_ficha(request, inscripcion_id):
    inscripcion = get_object_or_404(Inscripcion, id=inscripcion_id)

    if request.user.rol not in ["ADMINISTRADOR", "VALIDADOR"]:
        raise PermissionDenied

    try:
        pdf_bytes = generar_ficha_postulante_pdf(inscripcion)
        enviar_ficha_postulante(inscripcion, pdf_bytes)

        inscripcion.correo_enviado = True
        inscripcion.save(update_fields=["correo_enviado"])

        messages.success(
            request,
            "La ficha fue reenviada correctamente."
        )

    except Exception as e:
        logger.error(
            f"Error reenviando ficha {inscripcion.id}: {str(e)}"
        )
        messages.error(
            request,
            "No se pudo reenviar la ficha en este momento."
        )

    return redirect("panel:dashboard")



@usuario_interno_requerido
def exportar_postulantes_excel(request):
    estado = request.GET.get("estado", "").strip()
    q = request.GET.get("q", "").strip()

    inscripciones = Inscripcion.objects.select_related(
        "postulante",
        "convocatoria",
        "modalidad",
        "primera_opcion_programa",
        "segunda_opcion_programa",
    ).order_by("-fecha_creacion")

    if estado:
        inscripciones = inscripciones.filter(estado=estado)

    if q:
        inscripciones = inscripciones.filter(
            postulante__nombres__icontains=q
        ) | inscripciones.filter(
            postulante__apellido_paterno__icontains=q
        ) | inscripciones.filter(
            postulante__apellido_materno__icontains=q
        ) | inscripciones.filter(
            postulante__numero_documento__icontains=q
        ) | inscripciones.filter(
            numero_inscripcion__icontains=q
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Postulantes"

    encabezados = [
        "N° Inscripción",
        "Código Postulante",
        "Convocatoria",
        "Año",
        "Modalidad",
        "Estado",
        "Apellido Paterno",
        "Apellido Materno",
        "Nombres",
        "Tipo Documento",
        "Número Documento",
        "Sexo",
        "Correo",
        "Celular",
        "Idioma Materno",
        "Fecha Nacimiento",
        "Lugar Nacimiento",
        "Distrito Nacimiento",
        "Provincia Nacimiento",
        "Departamento Nacimiento",
        "País Nacimiento",
        "Institución Procedencia",
        "Año Egreso",
        "Gestión Institución",
        "Dirección Institución",
        "Distrito Institución",
        "Provincia Institución",
        "Departamento Institución",
        "País Institución",
        "Es Menor de Edad",
        "Tutor Apellidos",
        "Tutor Nombres",
        "Tutor Documento",
        "Tutor Parentesco",
        "Primera Opción",
        "Segunda Opción",
        "Medio Información",
        "Código Voucher",
        "Fecha Registro",
        "Fecha Validación",
        "Observaciones",
    ]

    ws.append(encabezados)

    fill = PatternFill("solid", fgColor="1D4ED8")
    font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for inscripcion in inscripciones:
        postulante = inscripcion.postulante

        idioma = postulante.get_idioma_materno_display()
        if postulante.idioma_materno == "OTRO" and postulante.idioma_materno_otro:
            idioma = f"{idioma} - {postulante.idioma_materno_otro}"

        lugar_nacimiento = postulante.lugar_nacimiento or ""
        distrito_nacimiento = postulante.distrito_nacimiento or ""
        provincia_nacimiento = postulante.provincia_nacimiento or ""
        departamento_nacimiento = postulante.departamento_nacimiento or ""
        pais_nacimiento = postulante.pais_nacimiento or ""

        segunda_opcion = (
            inscripcion.segunda_opcion_programa.nombre
            if inscripcion.segunda_opcion_programa else ""
        )

        ws.append([
            inscripcion.numero_inscripcion,
            inscripcion.codigo_postulante,
            inscripcion.convocatoria.nombre,
            inscripcion.convocatoria.anio,
            inscripcion.modalidad.nombre,
            inscripcion.estado,
            postulante.apellido_paterno,
            postulante.apellido_materno,
            postulante.nombres,
            postulante.tipo_documento,
            postulante.numero_documento,
            postulante.get_sexo_display(),
            postulante.correo_electronico,
            postulante.celular,
            idioma,
            postulante.fecha_nacimiento.strftime("%d/%m/%Y") if postulante.fecha_nacimiento else "",
            lugar_nacimiento,
            distrito_nacimiento,
            provincia_nacimiento,
            departamento_nacimiento,
            pais_nacimiento,
            postulante.institucion_procedencia,
            postulante.anio_egreso,
            postulante.get_gestion_institucion_display() if postulante.gestion_institucion else "",
            postulante.direccion_institucion,
            postulante.distrito_institucion,
            postulante.provincia_institucion,
            postulante.departamento_institucion,
            postulante.pais_institucion,
            "Sí" if postulante.es_menor_edad else "No",
            postulante.tutor_apellidos,
            postulante.tutor_nombres,
            postulante.tutor_numero_documento,
            postulante.tutor_tipo_parentesco,
            inscripcion.primera_opcion_programa.nombre,
            segunda_opcion,
            inscripcion.medio_informacion_admision,
            inscripcion.codigo_voucher_pago,
            inscripcion.fecha_creacion.strftime("%d/%m/%Y %H:%M") if inscripcion.fecha_creacion else "",
            inscripcion.fecha_validacion.strftime("%d/%m/%Y %H:%M") if inscripcion.fecha_validacion else "",
            inscripcion.observaciones_generales,
        ])

    anchos = {
        "A": 20, "B": 22, "C": 18, "D": 10, "E": 18, "F": 14,
        "G": 18, "H": 18, "I": 22, "J": 18, "K": 18, "L": 12,
        "M": 28, "N": 16, "O": 20, "P": 16, "Q": 20, "R": 18,
        "S": 18, "T": 20, "U": 14, "V": 24, "W": 12, "X": 18,
        "Y": 26, "Z": 18, "AA": 18, "AB": 18, "AC": 14, "AD": 20,
        "AE": 20, "AF": 18, "AG": 18, "AH": 20, "AI": 20, "AJ": 20,
        "AK": 18, "AL": 18, "AM": 24
    }

    for col, width in anchos.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="postulantes_admision.xlsx"'

    wb.save(response)
    return response